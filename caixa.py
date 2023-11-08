import openpyxl
import tkinter as tk
from datetime import datetime
import os

# Defina o diretório de trabalho para "C:\Users\User\Desktop\Caixa Loja"
diretorio_trabalho = r'C:\Users\User\Desktop\Caixa Loja'
os.chdir(diretorio_trabalho)

# Verifique o diretório de trabalho atual
print(f"Diretório de Trabalho Atual: {os.getcwd()}")

# Verifique se o arquivo Excel já existe, e se sim, carregue-o
nome_arquivo_excel = "novo_registro_vendas.xlsx"
if os.path.isfile(nome_arquivo_excel):
    workbook = openpyxl.load_workbook(nome_arquivo_excel)
else:
    # Se o arquivo não existe, crie um novo workbook
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove a planilha padrão

# Função para adicionar um produto à venda
def adicionar_produto():
    produto = produto_entry.get()
    preco = float(preco_entry.get())
    quantidade = int(quantidade_entry.get())
    metodo_pagamento = metodo_pagamento_var.get()

    # Adicionar o produto à lista de vendas
    venda = {
        "produto": produto,
        "preco": preco,
        "quantidade": quantidade,
        "metodo_pagamento": metodo_pagamento
    }
    lista_vendas.append(venda)

    # Atualizar a lista de produtos adicionados à venda
    lista_produtos.insert(tk.END, f"{quantidade} x {produto} ({metodo_pagamento}) - R$ {preco * quantidade:.2f}")

    # Limpar os campos de entrada
    produto_entry.delete(0, "end")
    preco_entry.delete(0, "end")
    quantidade_entry.delete(0, "end")

# Função para registrar a venda na planilha
def registrar_venda():
    # Selecionar a planilha ativa ou criar uma nova
    if "Vendas" in workbook.sheetnames:
        sheet = workbook["Vendas"]
    else:
        sheet = workbook.create_sheet("Vendas")
        sheet.append(["Data", "Produto", "Preço", "Quantidade", "Total Venda", "Método de Pagamento"])

    data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Registrar cada produto na planilha
    for venda in lista_vendas:
        produto = venda["produto"]
        preco = venda["preco"]
        quantidade = venda["quantidade"]
        metodo_pagamento = venda["metodo_pagamento"]
        total_venda = preco * quantidade
        sheet.append([data, produto, preco, quantidade, total_venda, metodo_pagamento])

    # Limpar a lista de vendas e a lista de produtos
    lista_vendas.clear()
    lista_produtos.delete(0, tk.END)

    # Salvar a planilha
    workbook.save(nome_arquivo_excel)
    resultado_label.config(text="Venda registrada com sucesso!", fg="green")

# Inicializar a interface gráfica
root = tk.Tk()
root.title("Gabbia Calçados - Caixa")

# Estilo da fonte
fonte = ("Arial", 12)

# Widgets
titulo_label = tk.Label(root, text="Registro de Vendas", font=("Arial", 16, "bold"))
titulo_label.pack(pady=10)

produto_label = tk.Label(root, text="Produto:", font=fonte)
produto_label.pack()
produto_entry = tk.Entry(root, font=fonte)
produto_entry.pack()

preco_label = tk.Label(root, text="Preço unitário:", font=fonte)
preco_label.pack()
preco_entry = tk.Entry(root, font=fonte)
preco_entry.pack()

quantidade_label = tk.Label(root, text="Quantidade:", font=fonte)
quantidade_label.pack()
quantidade_entry = tk.Entry(root, font=fonte)
quantidade_entry.pack()

metodo_pagamento_label = tk.Label(root, text="Método de pagamento:", font=fonte)
metodo_pagamento_label.pack()
metodo_pagamento_var = tk.StringVar()
metodo_pagamento_var.set("Pix")
metodo_pagamento_options = ["Pix", "Cartão", "Dinheiro"]
metodo_pagamento_menu = tk.OptionMenu(root, metodo_pagamento_var, *metodo_pagamento_options)
metodo_pagamento_menu.config(font=fonte)
metodo_pagamento_menu.pack()

adicionar_produto_button = tk.Button(root, text="Adicionar Produto", command=adicionar_produto, font=("Arial", 14, "bold"))
adicionar_produto_button.pack(pady=10)

lista_produtos = tk.Listbox(root, font=fonte)
lista_produtos.pack()

registrar_button = tk.Button(root, text="Registrar Venda", command=registrar_venda, font=("Arial", 14, "bold"))
registrar_button.pack(pady=10)

resultado_label = tk.Label(root, text="", font=("Arial", 14), fg="red")
resultado_label.pack()

# Lista para acompanhar todos os itens vendidos
lista_vendas = []

root.geometry("400x500")  # Tamanho da janela

root.mainloop()
